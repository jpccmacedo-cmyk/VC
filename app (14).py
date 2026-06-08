import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from copy import copy
from datetime import datetime, date, timedelta
from io import BytesIO
import zipfile
import tempfile
import shutil
import os
import re
import gc
import time
import unicodedata
import traceback
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

st.set_page_config(page_title="Consolidador Gerencial CN", page_icon="📊", layout="wide")

EXTENSOES_EXCEL_VALIDAS = (".xlsx", ".xlsm")
NOME_ABA_GERENCIAL = "Gerencial"
COLUNA_INICIO_LIMPEZA = 12  # Coluna L
ORDEM_ABAS = ["COB", "CUI", "EDE", "NOB", "PVE", "SOB", "XAM"]
MESES_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}


# ============================================================
# UTILITÁRIOS
# ============================================================

def remover_acentos(texto):
    texto = str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ASCII", "ignore").decode("ASCII")
    return texto


def normalizar_texto(texto):
    return str(texto).strip().lower()


def formatar_duracao(segundos):
    segundos = int(max(0, segundos))
    h = segundos // 3600
    m = (segundos % 3600) // 60
    s = segundos % 60
    if h:
        return f"{h}h {m:02d}min {s:02d}s"
    if m:
        return f"{m}min {s:02d}s"
    return f"{s}s"


def calcular_estimativa_tempo(inicio, concluidos, total):
    if concluidos <= 0 or total <= 0:
        return "calculando...", "calculando..."
    decorrido = time.time() - inicio
    restante = (decorrido / concluidos) * max(total - concluidos, 0)
    return formatar_duracao(decorrido), formatar_duracao(restante)


def limpar_nome_arquivo(nome):
    nome = re.sub(r'[\\/:*?"<>|]', '', str(nome)).strip()
    return nome if nome else "arquivo.xlsx"


def limpar_nome_aba(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = re.sub(r'[:\\/?*\[\]]', '', nome).strip()
    return (nome if nome else "Aba")[:31]


def gerar_nome_aba_unico(nome_base, nomes_existentes):
    nome_base = limpar_nome_aba(nome_base)
    if nome_base not in nomes_existentes:
        return nome_base
    contador = 1
    while True:
        sufixo = f" ({contador})"
        tentativa = nome_base[:31 - len(sufixo)] + sufixo
        if tentativa not in nomes_existentes:
            return tentativa
        contador += 1


def normalizar_nome_para_mapeamento(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = remover_acentos(nome).lower()
    nome = re.sub(r"[_\-.]+", " ", nome)
    nome = re.sub(r"\s+", " ", nome)
    return nome.strip()


def obter_nome_aba_final_personalizado(nome_arquivo_origem):
    nome_original_sem_ext = os.path.splitext(str(nome_arquivo_origem))[0]
    nome_sem_acento = remover_acentos(nome_original_sem_ext).lower().strip()
    nome_sem_acento = re.sub(r"\s+", " ", nome_sem_acento)
    nome_normalizado = normalizar_nome_para_mapeamento(nome_arquivo_origem)

    if "resumo gerencial cuiaba" in nome_normalizado:
        return "CUI"
    if "resumo gerencial diario ede" in nome_normalizado:
        return "EDE"
    if "rg sobradinho" in nome_normalizado:
        return "SOB"
    if re.search(r"(^|[^a-z0-9])gerencial_([0-2]?\d|3[01])([^0-9]|$)", nome_sem_acento):
        return "XAM"
    if re.search(r"\bgerencial\s+([0-2]\d|3[01])-[01]\d-\d{4}\b", nome_sem_acento):
        return "COB"
    if re.search(r"\bgerencial\s+([0-2]?\d|3[01])([^0-9-]|$)", nome_sem_acento):
        return "PVE"
    if "resumo gerencial" in nome_normalizado:
        return "NOB"
    return limpar_nome_aba(nome_arquivo_origem)


def identificar_aba_padrao(nome_aba):
    nome = limpar_nome_aba(nome_aba).upper().strip()
    for prefixo in ORDEM_ABAS:
        if nome == prefixo or nome.startswith(prefixo + " ") or nome.startswith(prefixo + "_") or nome.startswith(prefixo + "("):
            return prefixo
    return None


def aba_padrao_ja_existe(wb_final, nome_padrao):
    return any(identificar_aba_padrao(ws.title) == nome_padrao for ws in wb_final.worksheets)


# ============================================================
# LEITURA DE DATA EXCLUSIVAMENTE EM J5/K5
# ============================================================

def converter_para_date(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def interpretar_data_em_texto(texto):
    if texto is None:
        return None
    texto = str(texto).strip()
    if not texto:
        return None

    t = remover_acentos(texto).lower().strip()
    t = re.sub(r"\s+", " ", t)

    # Ex.: 3 de junho de 2026 / 03 de junho de 2026
    m = re.search(
        r"\b([0-3]?\d)\s+de\s+(janeiro|fevereiro|marco|março|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s+de\s+(\d{4})\b",
        t,
    )
    if m:
        try:
            dia = int(m.group(1))
            mes_nome = m.group(2)
            mes = MESES_PT.get(mes_nome, MESES_PT.get(remover_acentos(mes_nome)))
            ano = int(m.group(3))
            if mes:
                return datetime(ano, mes, dia).date()
        except ValueError:
            pass

    # Ex.: 03/06/2026, 03.06.2026, 03-06-2026
    m = re.search(r"\b([0-3]?\d)[\.\-/]([01]?\d)[\.\-/](\d{4})\b", t)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
        except ValueError:
            pass

    return None


def extrair_data_da_aba_pelo_conteudo(aba):
    """
    LÓGICA ÚNICA DE DATA:
    - O sistema lê SOMENTE as células J5 e K5.
    - Não verifica nome da aba.
    - Não procura em outras linhas/colunas.
    - Exemplo esperado: J5 = 'Data:' e K5 = '3 de junho de 2026'.
    """
    valores = []
    for ref in ["J5", "K5"]:
        try:
            valor = aba[ref].value
        except Exception:
            valor = None

        data_valor = converter_para_date(valor)
        if data_valor:
            return data_valor

        if valor is not None:
            valores.append(str(valor))

    texto_j5_k5 = " ".join(valores).strip()
    return interpretar_data_em_texto(texto_j5_k5)


def obter_data_real_da_aba(aba, ano_referencia=None, mes_referencia=None):
    """
    Retorna somente a data encontrada em J5/K5.
    Não usa nome da aba como fallback.
    """
    return extrair_data_da_aba_pelo_conteudo(aba)


def encontrar_aba_gerencial(wb_origem):
    for nome_aba in wb_origem.sheetnames:
        if normalizar_texto(nome_aba) == normalizar_texto(NOME_ABA_GERENCIAL):
            return nome_aba
    return None


def encontrar_aba_exata_por_data(wb_origem, data_selecionada, ano_referencia=None, mes_referencia=None):
    """
    Busca somente abas cuja data em J5/K5 seja igual à data selecionada.
    Não considera nome de aba.
    """
    candidatas = []
    for nome_aba in wb_origem.sheetnames:
        aba = wb_origem[nome_aba]
        data_real = obter_data_real_da_aba(aba)
        if data_real == data_selecionada:
            prioridade = 2 if normalizar_texto(nome_aba) == normalizar_texto(NOME_ABA_GERENCIAL) else 1
            candidatas.append({"nome_aba": nome_aba, "prioridade": prioridade})

    if candidatas:
        candidatas = sorted(candidatas, key=lambda item: item["prioridade"])
        return candidatas[0]["nome_aba"], "Aba com data real lida exclusivamente em J5/K5"

    return None, "Nenhuma aba com data real em J5/K5 igual à data selecionada encontrada"


def encontrar_aba_por_data_ou_gerencial(wb_origem, data_selecionada, ano_referencia=None, mes_referencia=None, permitir_fallback_gerencial=False):
    nome_aba, criterio = encontrar_aba_exata_por_data(wb_origem, data_selecionada)
    if nome_aba:
        return nome_aba, criterio

    if permitir_fallback_gerencial:
        aba_gerencial = encontrar_aba_gerencial(wb_origem)
        if aba_gerencial:
            # Mesmo com fallback, valida se J5/K5 da aba Gerencial bate com a data.
            data_gerencial = obter_data_real_da_aba(wb_origem[aba_gerencial])
            if data_gerencial == data_selecionada:
                return aba_gerencial, "Aba Gerencial usada por opção do usuário e validada por J5/K5"

    return None, "Nenhuma aba com data real em J5/K5 igual à data selecionada encontrada"


def listar_datas_disponiveis_arquivos(arquivos_salvos, ano_referencia, mes_referencia, quantidade_datas_filtro=10):
    """
    Lista datas disponíveis lendo exclusivamente J5/K5 de cada aba.
    Não considera nome da aba.
    """
    datas_encontradas = {}
    for item in arquivos_salvos:
        wb = None
        try:
            if not os.path.exists(item["caminho"]):
                continue
            wb = load_workbook(item["caminho"], read_only=False, data_only=True, keep_links=False)
            for nome_aba in wb.sheetnames:
                aba = wb[nome_aba]
                data_real = obter_data_real_da_aba(aba)
                if data_real and data_real.year == ano_referencia and data_real.month == mes_referencia:
                    datas_encontradas.setdefault(
                        data_real,
                        {"data": datetime.combine(data_real, datetime.min.time()), "abas": [], "arquivos": []},
                    )
                    datas_encontradas[data_real]["abas"].append(nome_aba)
                    datas_encontradas[data_real]["arquivos"].append(item["nome"])
        except Exception:
            pass
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            gc.collect()

    if not datas_encontradas:
        return []
    lista_datas = sorted(datas_encontradas.values(), key=lambda item: item["data"], reverse=True)
    return lista_datas[:quantidade_datas_filtro]


def extrair_data_nome_consolidado(nome_arquivo):
    """
    Esta função só é usada para impedir que consolidados antigos com data no nome
    entrem na data errada. A seleção das abas continua sendo somente por J5/K5.
    """
    m = re.search(r"(\d{2})[\.\-/](\d{2})[\.\-/](\d{4})", str(nome_arquivo))
    if not m:
        return None
    try:
        return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date()
    except ValueError:
        return None


# ============================================================
# CONSOLIDADO EXISTENTE
# ============================================================

def detectar_consolidado_existente(item):
    nome = remover_acentos(item["nome"]).lower()
    if "resumo gerencial cn" in nome or "consolidado" in nome:
        return True
    wb = None
    try:
        wb = load_workbook(item["caminho"], read_only=True, data_only=True, keep_links=False)
        return sum(1 for aba in wb.sheetnames if identificar_aba_padrao(aba) is not None) >= 2
    except Exception:
        return False
    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass


def consolidado_serve_para_data(item, data_selecionada, total_datas_selecionadas):
    data_nome = extrair_data_nome_consolidado(item["nome"])
    if data_nome is not None:
        return data_nome == data_selecionada
    return total_datas_selecionadas == 1


def copiar_abas_de_consolidado_existente(wb_origem, wb_final, logs, nome_arquivo, data_selecionada):
    copiadas = []
    for nome_aba in wb_origem.sheetnames:
        nome_padrao = identificar_aba_padrao(nome_aba)
        if nome_padrao is None:
            continue

        aba_origem = wb_origem[nome_aba]
        data_real = obter_data_real_da_aba(aba_origem)
        if data_real != data_selecionada:
            logs.append(
                f"Consolidado existente: aba '{nome_aba}' ignorada porque J5/K5 indica "
                f"{data_real.strftime('%d/%m/%Y') if data_real else 'sem data'}, não {data_selecionada.strftime('%d/%m/%Y')}."
            )
            continue

        if aba_padrao_ja_existe(wb_final, nome_padrao):
            logs.append(f"Consolidado existente: aba '{nome_aba}' ignorada porque '{nome_padrao}' já existe.")
            continue

        nome_destino = gerar_nome_aba_unico(nome_padrao, wb_final.sheetnames)
        aba_destino = wb_final.create_sheet(title=nome_destino)
        copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=True)
        limpar_conteudos_a_partir_da_coluna_l(aba_destino)
        aplicar_configuracoes_finais_aba(aba_destino)
        copiadas.append(nome_destino)
        logs.append(f"Consolidado existente '{nome_arquivo}': aba '{nome_aba}' copiada como '{nome_destino}'.")
    return copiadas


# ============================================================
# UPLOAD / SESSÃO
# ============================================================

def criar_pasta_sessao():
    if "pasta_sessao" not in st.session_state:
        st.session_state.pasta_sessao = tempfile.mkdtemp()


def salvar_uploads_em_pasta_sessao(arquivos_upload):
    criar_pasta_sessao()
    arquivos_salvos = []
    for uploaded_file in arquivos_upload:
        nome_limpo = limpar_nome_arquivo(uploaded_file.name)
        caminho_base = os.path.join(st.session_state.pasta_sessao, nome_limpo)
        caminho_final = caminho_base
        contador = 1
        while os.path.exists(caminho_final):
            nome_sem_extensao, extensao = os.path.splitext(nome_limpo)
            caminho_final = os.path.join(st.session_state.pasta_sessao, f"{nome_sem_extensao}_{contador}{extensao}")
            contador += 1
        with open(caminho_final, "wb") as f:
            f.write(uploaded_file.getvalue())
        arquivos_salvos.append({"nome": uploaded_file.name, "caminho": caminho_final, "tamanho": uploaded_file.size})
    return arquivos_salvos


def limpar_arquivos_da_sessao():
    if "pasta_sessao" in st.session_state:
        try:
            if os.path.exists(st.session_state.pasta_sessao):
                shutil.rmtree(st.session_state.pasta_sessao)
        except Exception:
            pass
    st.session_state.arquivos_salvos = []
    st.session_state.uploader_key += 1
    if "pasta_sessao" in st.session_state:
        del st.session_state.pasta_sessao


# ============================================================
# FORMATAÇÃO / CÓPIA
# ============================================================

def aplicar_configuracoes_finais_aba(aba_destino):
    aba_destino.sheet_view.zoomScale = 80
    aba_destino.sheet_view.zoomScaleNormal = 80
    aba_destino.freeze_panes = "A5"
    aba_destino.sheet_view.showGridLines = False


def limpar_conteudos_a_partir_da_coluna_l(aba_destino):
    if aba_destino.max_column < COLUNA_INICIO_LIMPEZA:
        return
    for intervalo in list(aba_destino.merged_cells.ranges):
        if intervalo.min_col >= COLUNA_INICIO_LIMPEZA:
            cel = aba_destino.cell(row=intervalo.min_row, column=intervalo.min_col)
            if not isinstance(cel, MergedCell):
                cel.value = None
                cel.comment = None
                cel._hyperlink = None
    for linha in aba_destino.iter_rows(min_row=1, max_row=aba_destino.max_row, min_col=COLUNA_INICIO_LIMPEZA, max_col=aba_destino.max_column):
        for cel in linha:
            if isinstance(cel, MergedCell):
                continue
            cel.value = None
            cel.comment = None
            cel._hyperlink = None


def garantir_requisitos_todas_abas(wb_final):
    for aba in wb_final.worksheets:
        limpar_conteudos_a_partir_da_coluna_l(aba)
        aplicar_configuracoes_finais_aba(aba)


def padronizar_larguras_colunas_pela_aba_ede(wb_final, nome_aba_referencia="EDE"):
    if nome_aba_referencia not in wb_final.sheetnames:
        return False, f"A aba de referência '{nome_aba_referencia}' não foi encontrada. As larguras não foram padronizadas."
    aba_referencia = wb_final[nome_aba_referencia]
    max_colunas = max(ws.max_column for ws in wb_final.worksheets)
    for idx in range(1, max_colunas + 1):
        letra = aba_referencia.cell(row=1, column=idx).column_letter
        largura = aba_referencia.column_dimensions[letra].width
        if largura is None:
            largura = aba_referencia.sheet_format.defaultColWidth or 8.43
        for aba in wb_final.worksheets:
            aba.column_dimensions[letra].width = largura
    return True, f"Larguras padronizadas com base na aba '{nome_aba_referencia}'."


def copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=True):
    for linha in aba_origem.iter_rows(min_row=1, max_row=aba_origem.max_row, min_col=1, max_col=aba_origem.max_column):
        for celula_origem in linha:
            celula_destino = aba_destino[celula_origem.coordinate]
            celula_destino.value = celula_origem.value
            if copiar_estilos and celula_origem.has_style:
                celula_destino.font = copy(celula_origem.font)
                celula_destino.fill = copy(celula_origem.fill)
                celula_destino.border = copy(celula_origem.border)
                celula_destino.alignment = copy(celula_origem.alignment)
                celula_destino.number_format = celula_origem.number_format
                celula_destino.protection = copy(celula_origem.protection)
            if celula_origem.comment:
                celula_destino.comment = copy(celula_origem.comment)
            if celula_origem.hyperlink:
                celula_destino._hyperlink = copy(celula_origem.hyperlink)
    for letra, dim in aba_origem.column_dimensions.items():
        aba_destino.column_dimensions[letra].width = dim.width
        aba_destino.column_dimensions[letra].hidden = dim.hidden
        aba_destino.column_dimensions[letra].outlineLevel = dim.outlineLevel
        aba_destino.column_dimensions[letra].collapsed = dim.collapsed
    for numero, dim in aba_origem.row_dimensions.items():
        aba_destino.row_dimensions[numero].height = dim.height
        aba_destino.row_dimensions[numero].hidden = dim.hidden
        aba_destino.row_dimensions[numero].outlineLevel = dim.outlineLevel
        aba_destino.row_dimensions[numero].collapsed = dim.collapsed
    for intervalo in aba_origem.merged_cells.ranges:
        aba_destino.merge_cells(str(intervalo))
    if aba_origem.auto_filter and aba_origem.auto_filter.ref:
        aba_destino.auto_filter.ref = aba_origem.auto_filter.ref


def ordenar_abas(wb_final):
    def prioridade(aba):
        nome = aba.title.upper()
        for idx, prefixo in enumerate(ORDEM_ABAS):
            if nome == prefixo or nome.startswith(prefixo + " ") or nome.startswith(prefixo + "_") or nome.startswith(prefixo + "("):
                return idx
        return 99
    wb_final._sheets = sorted(wb_final.worksheets, key=prioridade)


# ============================================================
# PROCESSAMENTO
# ============================================================

def processar_arquivos_salvos(arquivos_salvos, data_referencia, data_selecionada, copiar_estilos=True, callback_progresso=None, total_datas_selecionadas=1, permitir_fallback_gerencial=False):
    logs, arquivos_processados, arquivos_com_erro, abas_criadas = [], [], [], []
    nome_arquivo_final = f"Resumo Gerencial CN - {data_selecionada.strftime('%d.%m.%Y')}.xlsx"
    wb_final = Workbook()
    wb_final.remove(wb_final.active)

    logs.append(f"Data do consolidado: {data_selecionada.strftime('%d/%m/%Y')}")
    logs.append("Validação única: data lida exclusivamente em J5/K5. Nome da aba não é considerado.")

    itens_consolidados, itens_origem = [], []
    for item in arquivos_salvos:
        (itens_consolidados if detectar_consolidado_existente(item) else itens_origem).append(item)

    for item in itens_consolidados:
        wb = None
        try:
            if not consolidado_serve_para_data(item, data_selecionada, total_datas_selecionadas):
                logs.append(f"Consolidado existente '{item['nome']}' ignorado nesta data.")
                continue
            wb = load_workbook(item["caminho"], data_only=True, keep_links=False)
            copiadas = copiar_abas_de_consolidado_existente(wb, wb_final, logs, item["nome"], data_selecionada)
            abas_criadas.extend(copiadas)
            if copiadas:
                arquivos_processados.append({
                    "arquivo": item["nome"],
                    "aba_origem": "Consolidado existente",
                    "aba_final": ", ".join(copiadas),
                    "criterio": "Abas reaproveitadas de consolidado existente com J5/K5 validado"
                })
        except Exception as erro:
            arquivos_com_erro.append({"arquivo": item["nome"], "erro": str(erro)})
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass

    for item in itens_origem:
        wb = None
        try:
            nome_original = item["nome"]
            nome_base = obter_nome_aba_final_personalizado(nome_original)
            nome_padrao = identificar_aba_padrao(nome_base) or nome_base

            if aba_padrao_ja_existe(wb_final, nome_padrao):
                logs.append(f"Arquivo '{nome_original}' ignorado porque '{nome_padrao}' já existe.")
                if callback_progresso:
                    callback_progresso(nome_original, data_selecionada)
                continue

            wb = load_workbook(item["caminho"], data_only=True, keep_links=False)
            nome_aba_origem, criterio = encontrar_aba_por_data_ou_gerencial(wb, data_selecionada, permitir_fallback_gerencial=permitir_fallback_gerencial)

            if not nome_aba_origem:
                logs.append(f"IGNORADO: {nome_original} não possui data real {data_selecionada.strftime('%d/%m/%Y')} em J5/K5.")
                if callback_progresso:
                    callback_progresso(nome_original, data_selecionada)
                continue

            aba_origem = wb[nome_aba_origem]
            data_real = obter_data_real_da_aba(aba_origem)
            if data_real != data_selecionada:
                logs.append(f"IGNORADO: {nome_original} tem J5/K5 = {data_real.strftime('%d/%m/%Y') if data_real else 'sem data'}, não {data_selecionada.strftime('%d/%m/%Y')}.")
                if callback_progresso:
                    callback_progresso(nome_original, data_selecionada)
                continue

            nome_final = gerar_nome_aba_unico(nome_base, wb_final.sheetnames)
            aba_destino = wb_final.create_sheet(title=nome_final)
            copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=copiar_estilos)
            limpar_conteudos_a_partir_da_coluna_l(aba_destino)
            aplicar_configuracoes_finais_aba(aba_destino)

            arquivos_processados.append({"arquivo": nome_original, "aba_origem": nome_aba_origem, "aba_final": nome_final, "criterio": criterio})
            abas_criadas.append(nome_final)
            logs.append(f"SUCESSO: {nome_original} | aba '{nome_aba_origem}' -> '{nome_final}'.")

        except Exception as erro:
            arquivos_com_erro.append({"arquivo": item.get("nome", "arquivo"), "erro": str(erro)})
            logs.append(f"ERRO: {item.get('nome', 'arquivo')} - {erro}")
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            gc.collect()
            if callback_progresso:
                callback_progresso(item.get("nome", "arquivo"), data_selecionada)

    if not wb_final.sheetnames:
        raise Exception(f"Nenhuma aba com data real {data_selecionada.strftime('%d/%m/%Y')} em J5/K5 foi criada.")

    garantir_requisitos_todas_abas(wb_final)
    _, msg = padronizar_larguras_colunas_pela_aba_ede(wb_final, "EDE")
    logs.append(msg)
    garantir_requisitos_todas_abas(wb_final)
    ordenar_abas(wb_final)

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    wb_final.save(temp)
    wb_final.close()
    with open(temp, "rb") as f:
        conteudo = f.read()
    try:
        os.remove(temp)
    except Exception:
        pass

    output = BytesIO(conteudo)
    output.seek(0)
    return {
        "arquivo_excel": output,
        "nome_arquivo_final": nome_arquivo_final,
        "logs": logs,
        "arquivos_processados": arquivos_processados,
        "arquivos_com_erro": arquivos_com_erro,
        "abas_criadas": abas_criadas,
        "data_processada": data_selecionada,
    }


def processar_multiplas_datas(arquivos_salvos, datas_selecionadas, data_referencia, copiar_estilos=True, permitir_fallback_gerencial=False):
    resultados = []
    datas_ordenadas = sorted(datas_selecionadas, key=lambda x: x["data"], reverse=True)
    total_etapas = max(len(datas_ordenadas) * len(arquivos_salvos), 1)
    etapas = 0
    inicio = time.time()
    barra = st.progress(0)
    status = st.empty()
    tempo_box = st.empty()

    def progresso(nome_arquivo, data_atual):
        nonlocal etapas
        etapas += 1
        barra.progress(min(etapas / total_etapas, 1))
        decorrido, restante = calcular_estimativa_tempo(inicio, etapas, total_etapas)
        tempo_box.info(f"⏱️ Tempo decorrido: {decorrido} | Tempo restante estimado: {restante} | Progresso: {etapas}/{total_etapas} etapas")
        status.info(f"Processando data {data_atual.strftime('%d/%m/%Y')} | Arquivo: {nome_arquivo}")

    for i, item_data in enumerate(datas_ordenadas, start=1):
        data = item_data["data"]
        status.info(f"Gerando consolidado {i}/{len(datas_ordenadas)} somente com a data {data.strftime('%d/%m/%Y')}")
        resultado = processar_arquivos_salvos(
            arquivos_salvos,
            data_referencia,
            data,
            copiar_estilos=copiar_estilos,
            callback_progresso=progresso,
            total_datas_selecionadas=len(datas_ordenadas),
            permitir_fallback_gerencial=permitir_fallback_gerencial
        )
        resultados.append(resultado)

    decorrido, _ = calcular_estimativa_tempo(inicio, total_etapas, total_etapas)
    status.success("Todos os consolidados por data foram gerados.")
    tempo_box.success(f"✅ Processamento concluído. Tempo total: {decorrido}.")
    return resultados


def gerar_zip_resultados(resultados):
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for r in resultados:
            z.writestr(r["nome_arquivo_final"], r["arquivo_excel"].getvalue())
    zip_buffer.seek(0)
    return zip_buffer


# ============================================================
# ESTADO / INTERFACE
# ============================================================

if "arquivos_salvos" not in st.session_state:
    st.session_state.arquivos_salvos = []
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

st.title("📊 Sistema de Consolidação Gerencial CN")
st.markdown(
    """
    Este sistema gera **um consolidado separado para cada data selecionada**.

    A validação da data segue **somente uma regra**: ler a data nas células **J5/K5** da aba.
    O sistema **não usa o nome da aba** para identificar a data.
    """
)
st.divider()

st.sidebar.header("⚙️ Configurações")
usar_data_ontem = st.sidebar.checkbox("Usar mês/ano de ontem como referência", value=True)
if usar_data_ontem:
    data_referencia = datetime.today() - timedelta(days=1)
    st.sidebar.info(f"Mês/ano de referência: {data_referencia.strftime('%m/%Y')}")
else:
    data_escolhida = st.sidebar.date_input("Escolha uma data de referência para mês/ano", value=(datetime.today() - timedelta(days=1)).date())
    data_referencia = datetime.combine(data_escolhida, datetime.min.time())

qtde_esperada = st.sidebar.number_input("Quantidade esperada de arquivos", min_value=1, max_value=100, value=7, step=1)
copiar_estilos = st.sidebar.checkbox("Copiar formatação das células", value=True)
quantidade_datas_filtro = st.sidebar.number_input("Quantidade máxima de datas no filtro", min_value=1, max_value=31, value=5, step=1)
max_datas_processar = st.sidebar.number_input("Máximo de datas para processar de uma vez", min_value=1, max_value=10, value=5, step=1)
permitir_fallback_gerencial = st.sidebar.checkbox(
    "Permitir usar aba Gerencial quando J5/K5 bater com a data",
    value=False,
    help="Mesmo marcado, a aba Gerencial só será usada se J5/K5 tiver a mesma data selecionada."
)

st.subheader("1. Envie os arquivos Excel")
arquivos_upload = st.file_uploader(
    "Selecione todos os arquivos Excel de uma vez. Você pode incluir consolidados já gerados e planilhas faltantes.",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
    key=f"uploader_{st.session_state.uploader_key}"
)

col_carregar, col_limpar = st.columns(2)
with col_carregar:
    if st.button("📥 Carregar arquivos enviados", disabled=not arquivos_upload):
        try:
            limpar_arquivos_da_sessao()
            st.session_state.arquivos_salvos = salvar_uploads_em_pasta_sessao(arquivos_upload)
            st.success(f"{len(st.session_state.arquivos_salvos)} arquivo(s) carregado(s) com sucesso.")
            st.rerun()
        except Exception as erro:
            st.error(f"Erro ao carregar arquivos: {erro}")
with col_limpar:
    if st.button("🧹 Limpar arquivos carregados"):
        limpar_arquivos_da_sessao()
        st.success("Arquivos carregados foram limpos.")
        st.rerun()

arquivos_salvos = st.session_state.arquivos_salvos
if arquivos_salvos:
    st.success(f"{len(arquivos_salvos)} arquivo(s) carregado(s).")
    tamanho_total_mb = sum(item["tamanho"] for item in arquivos_salvos) / (1024 * 1024)
    st.info(f"Tamanho total carregado: {tamanho_total_mb:.2f} MB")
    with st.expander("Ver arquivos carregados"):
        for i, item in enumerate(arquivos_salvos, start=1):
            tamanho_mb = item["tamanho"] / (1024 * 1024)
            nome_padronizado = obter_nome_aba_final_personalizado(item["nome"])
            tipo = "Consolidado existente" if detectar_consolidado_existente(item) else "Planilha de origem/faltante"
            st.write(f"{i}. {item['nome']} — {tamanho_mb:.2f} MB — tipo: {tipo} — aba prevista: {nome_padronizado}")
else:
    st.info("Envie os arquivos Excel e clique em 'Carregar arquivos enviados'.")

st.divider()
st.subheader("2. Escolha uma ou mais datas para gerar consolidados separados")
datas_selecionadas = []

if arquivos_salvos:
    datas_para_filtro = listar_datas_disponiveis_arquivos(
        arquivos_salvos,
        data_referencia.year,
        data_referencia.month,
        quantidade_datas_filtro
    )
    if datas_para_filtro:
        opcoes_datas = []
        for item in datas_para_filtro:
            data_item = item["data"].date()
            qtd_arquivos = len(set(item["arquivos"]))
            label = f"{data_item.strftime('%d/%m/%Y')} — encontrada em {qtd_arquivos} arquivo(s)"
            opcoes_datas.append({"label": label, "data": data_item})

        datas_selecionadas = st.multiselect(
            "Selecione uma ou mais datas",
            options=opcoes_datas,
            default=opcoes_datas[:1],
            format_func=lambda item: item["label"]
        )
        if datas_selecionadas:
            st.success(f"{len(datas_selecionadas)} data(s) selecionada(s).")
            st.info("Será gerado um arquivo Excel consolidado separado para cada data selecionada.")
    else:
        st.warning("Nenhuma data foi encontrada em J5/K5 no mês/ano de referência.")
else:
    st.info("Carregue os arquivos para o sistema listar as datas disponíveis.")

st.divider()
st.subheader("3. Processar")

botao_processar = st.button(
    "🚀 Gerar consolidado(s) por data",
    type="primary",
    disabled=len(st.session_state.arquivos_salvos) == 0 or len(datas_selecionadas) == 0
)

if botao_processar:
    try:
        if len(datas_selecionadas) > max_datas_processar:
            st.warning(f"Selecione no máximo {max_datas_processar} data(s) por execução.")
            st.stop()

        with st.spinner("Gerando consolidado(s) separado(s) por data..."):
            resultados = processar_multiplas_datas(
                arquivos_salvos=st.session_state.arquivos_salvos,
                datas_selecionadas=datas_selecionadas,
                data_referencia=data_referencia,
                copiar_estilos=copiar_estilos,
                permitir_fallback_gerencial=permitir_fallback_gerencial
            )

        st.success("Consolidado(s) gerado(s) com sucesso!")
        total_sucesso = sum(len(r["arquivos_processados"]) for r in resultados)
        total_erros = sum(len(r["arquivos_com_erro"]) for r in resultados)
        col1, col2, col3 = st.columns(3)
        col1.metric("Arquivos finais gerados", len(resultados))
        col2.metric("Abas processadas com sucesso", total_sucesso)
        col3.metric("Erros", total_erros)

        st.subheader("4. Downloads")
        if len(resultados) > 1:
            zip_file = gerar_zip_resultados(resultados)
            st.download_button("⬇️ Baixar todos os consolidados em ZIP", zip_file, "Consolidados_Gerenciais_CN.zip", mime="application/zip")

        st.markdown("### Downloads individuais")
        for resultado in resultados:
            st.download_button(
                label=f"⬇️ Baixar {resultado['nome_arquivo_final']}",
                data=resultado["arquivo_excel"],
                file_name=resultado["nome_arquivo_final"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.subheader("5. Resumo")
        for resultado in resultados:
            data_proc = resultado["data_processada"].strftime("%d/%m/%Y")
            with st.expander(f"Resumo do consolidado da data {data_proc}"):
                if resultado["arquivos_processados"]:
                    st.markdown("#### ✅ Arquivos processados")
                    for item in resultado["arquivos_processados"]:
                        st.write(f"- **Arquivo:** {item['arquivo']} | **Aba origem:** {item['aba_origem']} | **Aba final:** {item['aba_final']} | **Critério:** {item['criterio']}")
                if resultado["arquivos_com_erro"]:
                    st.markdown("#### ❌ Arquivos com erro")
                    for item in resultado["arquivos_com_erro"]:
                        st.error(f"{item['arquivo']}: {item['erro']}")
                st.markdown("#### Logs")
                st.code("\n".join(resultado["logs"]), language="text")

    except Exception as erro:
        st.error(f"Erro geral ao processar os arquivos: {erro}")
        st.code(traceback.format_exc(), language="text")

st.divider()
st.caption(
    "Observação: cada data selecionada gera um arquivo consolidado separado. "
    "A data é validada exclusivamente nas células J5/K5. "
    "O sistema não usa o nome da aba para identificar data. "
    "Cada consolidado segue as regras de abas, layout, limpeza da coluna L e ordem COB, CUI, EDE, NOB, PVE, SOB, XAM."
)
