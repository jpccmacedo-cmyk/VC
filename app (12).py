import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell
from copy import copy
from datetime import datetime, date, timedelta
from io import BytesIO
import zipfile
import os
import re
import gc
import shutil
import tempfile
import traceback
import warnings
import unicodedata
import time

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

st.set_page_config(
    page_title="Consolidador Gerencial CN",
    page_icon="📊",
    layout="wide"
)

EXTENSOES_EXCEL_VALIDAS = (".xlsx", ".xlsm")
NOME_ABA_GERENCIAL = "Gerencial"
COLUNA_INICIO_LIMPEZA = 12  # Coluna L
ORDEM_ABAS = ["COB", "CUI", "EDE", "NOB", "PVE", "SOB", "XAM"]

MESES_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "março": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}


# ============================================================
# TEMPO / PROGRESSO
# ============================================================

def formatar_duracao(segundos):
    segundos = int(max(0, segundos))
    horas = segundos // 3600
    minutos = (segundos % 3600) // 60
    seg = segundos % 60

    if horas > 0:
        return f"{horas}h {minutos:02d}min {seg:02d}s"
    if minutos > 0:
        return f"{minutos}min {seg:02d}s"
    return f"{seg}s"


def calcular_estimativa_tempo(inicio, concluidos, total):
    if concluidos <= 0 or total <= 0:
        return "calculando...", "calculando..."

    decorrido = time.time() - inicio
    media_por_item = decorrido / concluidos
    restante = media_por_item * max(total - concluidos, 0)

    return formatar_duracao(decorrido), formatar_duracao(restante)


# ============================================================
# TEXTO / NOMES
# ============================================================

def remover_acentos(texto):
    texto = str(texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ASCII", "ignore").decode("ASCII")
    return texto


def normalizar_texto(texto):
    return str(texto).strip().lower()


def limpar_nome_arquivo(nome):
    nome = re.sub(r'[\\/:*?"<>|]', '', str(nome)).strip()
    return nome if nome else "arquivo.xlsx"


def limpar_nome_aba(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = re.sub(r'[:\\/?*\[\]]', '', nome).strip()
    return (nome if nome else "Aba")[:31]


def gerar_nome_aba_unico(nome_base, nomes_existentes):
    nome_base = limpar_nome_aba(nome_base)

    if nome_base not in nomes_existentes:
        return nome_base

    contador = 1
    while True:
        sufixo = f" ({contador})"
        nome_tentativa = nome_base[:31 - len(sufixo)] + sufixo
        if nome_tentativa not in nomes_existentes:
            return nome_tentativa
        contador += 1


def normalizar_nome_para_mapeamento(nome):
    nome = os.path.splitext(str(nome))[0]
    nome = remover_acentos(nome).lower()
    nome = re.sub(r"[_\-.]+", " ", nome)
    nome = re.sub(r"\s+", " ", nome)
    return nome.strip()


def obter_nome_aba_final_personalizado(nome_arquivo_origem):
    nome_original_sem_ext = os.path.splitext(str(nome_arquivo_origem))[0]
    nome_sem_acento = remover_acentos(nome_original_sem_ext).lower().strip()
    nome_sem_acento = re.sub(r"\s+", " ", nome_sem_acento)
    nome_normalizado = normalizar_nome_para_mapeamento(nome_arquivo_origem)

    if "resumo gerencial cuiaba" in nome_normalizado:
        return "CUI"
    if "resumo gerencial diario ede" in nome_normalizado:
        return "EDE"
    if "rg sobradinho" in nome_normalizado:
        return "SOB"
    if re.search(r"(^|[^a-z0-9])gerencial_([0-2]?\d|3[01])([^0-9]|$)", nome_sem_acento):
        return "XAM"
    if re.search(r"\bgerencial\s+([0-2]\d|3[01])-[01]\d-\d{4}\b", nome_sem_acento):
        return "COB"
    if re.search(r"\bgerencial\s+([0-2]?\d|3[01])([^0-9-]|$)", nome_sem_acento):
        return "PVE"
    if "resumo gerencial" in nome_normalizado:
        return "NOB"

    return limpar_nome_aba(nome_arquivo_origem)


def identificar_aba_padrao(nome_aba):
    nome = limpar_nome_aba(nome_aba).upper().strip()
    for prefixo in ORDEM_ABAS:
        if nome == prefixo or nome.startswith(prefixo + " ") or nome.startswith(prefixo + "_") or nome.startswith(prefixo + "("):
            return prefixo
    return None


def aba_padrao_ja_existe(wb_final, nome_padrao):
    for aba in wb_final.worksheets:
        if identificar_aba_padrao(aba.title) == nome_padrao:
            return True
    return False


# ============================================================
# DATAS / ABAS / CONTEÚDO DA PLANILHA
# ============================================================

def converter_para_date(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def interpretar_data_em_texto(texto):
    if texto is None:
        return None

    original = str(texto).strip()
    if not original:
        return None

    texto_sem_acento = remover_acentos(original).lower().strip()
    texto_sem_acento = re.sub(r"\s+", " ", texto_sem_acento)

    # Formato: 3 de junho de 2026 / 03 de junho de 2026
    padrao_extenso = re.search(
        r"\b([0-3]?\d)\s+de\s+(janeiro|fevereiro|marco|março|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s+de\s+(\d{4})\b",
        texto_sem_acento
    )
    if padrao_extenso:
        try:
            dia = int(padrao_extenso.group(1))
            mes_nome = padrao_extenso.group(2)
            ano = int(padrao_extenso.group(3))
            mes = MESES_PT.get(mes_nome, MESES_PT.get(remover_acentos(mes_nome)))
            if mes:
                return datetime(ano, mes, dia).date()
        except ValueError:
            pass

    # Formato: 03/06/2026, 03.06.2026, 03-06-2026
    padrao_numerico = re.search(r"\b([0-3]?\d)[\.\-/]([01]?\d)[\.\-/](\d{4})\b", texto_sem_acento)
    if padrao_numerico:
        try:
            dia = int(padrao_numerico.group(1))
            mes = int(padrao_numerico.group(2))
            ano = int(padrao_numerico.group(3))
            return datetime(ano, mes, dia).date()
        except ValueError:
            pass

    return None


def extrair_data_da_aba_pelo_conteudo(aba, max_linhas=15, max_colunas=20):
    """
    Lê a data dentro da própria aba, como no cabeçalho:
    Data: 3 de junho de 2026.

    Estratégia:
    1. Procura a palavra 'Data' nas primeiras linhas e lê células à direita.
    2. Procura qualquer texto com data por extenso ou numérica nas primeiras linhas.
    3. Aceita células de data do Excel.
    """
    max_linha = min(aba.max_row or 1, max_linhas)
    max_coluna = min(aba.max_column or 1, max_colunas)

    # 1) Prioridade para campo próximo ao rótulo Data:
    for linha in range(1, max_linha + 1):
        for coluna in range(1, max_coluna + 1):
            valor = aba.cell(row=linha, column=coluna).value
            valor_texto = remover_acentos(str(valor)).lower().strip() if valor is not None else ""

            if valor_texto.startswith("data") or valor_texto == "data:":
                # Procura na própria célula e até 8 células à direita.
                for col_busca in range(coluna, min(max_coluna, coluna + 8) + 1):
                    valor_busca = aba.cell(row=linha, column=col_busca).value
                    data_valor = converter_para_date(valor_busca)
                    if data_valor:
                        return data_valor
                    data_texto = interpretar_data_em_texto(valor_busca)
                    if data_texto:
                        return data_texto

    # 2) Varredura geral no topo da planilha.
    for linha in range(1, max_linha + 1):
        for coluna in range(1, max_coluna + 1):
            valor = aba.cell(row=linha, column=coluna).value
            data_valor = converter_para_date(valor)
            if data_valor:
                return data_valor
            data_texto = interpretar_data_em_texto(valor)
            if data_texto:
                return data_texto

    return None


def interpretar_aba_como_data(nome_aba, ano_referencia, mes_referencia):
    nome = str(nome_aba).strip()
    if not nome.isdigit():
        return None

    try:
        if len(nome) in [1, 2]:
            return datetime(ano_referencia, mes_referencia, int(nome))
        if len(nome) == 4:
            dia = int(nome[:2])
            mes = int(nome[2:])
            return datetime(ano_referencia, mes, dia)
        return None
    except ValueError:
        return None


def obter_data_real_da_aba(aba, ano_referencia, mes_referencia):
    """
    Data real da aba = data dentro do cabeçalho da planilha.
    Se não encontrar no conteúdo, usa o nome da aba como fallback de detecção.
    """
    data_conteudo = extrair_data_da_aba_pelo_conteudo(aba)
    if data_conteudo:
        return data_conteudo

    data_nome = interpretar_aba_como_data(aba.title, ano_referencia, mes_referencia)
    if data_nome:
        return data_nome.date()

    return None


def listar_datas_disponiveis_arquivos(arquivos_salvos, ano_referencia, mes_referencia, quantidade_datas_filtro=10):
    datas_encontradas = {}

    for item in arquivos_salvos:
        wb = None
        try:
            caminho = item["caminho"]
            if not os.path.exists(caminho):
                continue

            wb = load_workbook(
                filename=caminho,
                read_only=False,
                data_only=True,
                keep_links=False
            )

            for nome_aba in wb.sheetnames:
                aba = wb[nome_aba]
                data_real = obter_data_real_da_aba(aba, ano_referencia, mes_referencia)

                if data_real is not None:
                    if data_real.year != ano_referencia or data_real.month != mes_referencia:
                        continue

                    chave = data_real
                    datas_encontradas.setdefault(chave, {"data": datetime.combine(data_real, datetime.min.time()), "abas": [], "arquivos": []})
                    datas_encontradas[chave]["abas"].append(nome_aba)
                    datas_encontradas[chave]["arquivos"].append(item["nome"])

        except Exception:
            pass
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
            gc.collect()

    if not datas_encontradas:
        return []

    lista_datas = sorted(datas_encontradas.values(), key=lambda item: item["data"], reverse=True)

    sequencia = []
    data_anterior = None
    for item in lista_datas:
        data_atual = item["data"].date()
        if data_anterior is None:
            sequencia.append(item)
            data_anterior = data_atual
        elif (data_anterior - data_atual).days == 1:
            sequencia.append(item)
            data_anterior = data_atual
        else:
            break

        if len(sequencia) >= quantidade_datas_filtro:
            break

    return sequencia


def encontrar_aba_gerencial(wb_origem):
    for nome_aba in wb_origem.sheetnames:
        if normalizar_texto(nome_aba) == normalizar_texto(NOME_ABA_GERENCIAL):
            return nome_aba
    return None


def encontrar_aba_exata_por_data(wb_origem, data_selecionada, ano_referencia, mes_referencia):
    """
    Busca a aba cuja data REAL, lida no cabeçalho/conteúdo da planilha,
    seja igual à data selecionada.

    Isso evita que uma aba com nome 04, mas cabeçalho 03/06/2026,
    entre no consolidado de 04/06/2026.
    """
    candidatas = []

    for nome_aba in wb_origem.sheetnames:
        aba = wb_origem[nome_aba]
        data_real = obter_data_real_da_aba(aba, ano_referencia, mes_referencia)

        if data_real is not None and data_real == data_selecionada:
            data_nome = interpretar_aba_como_data(nome_aba, ano_referencia, mes_referencia)
            if data_nome is not None and data_nome.date() == data_selecionada:
                prioridade = 1
            elif normalizar_texto(nome_aba) == normalizar_texto(NOME_ABA_GERENCIAL):
                prioridade = 2
            else:
                prioridade = 3
            candidatas.append({"nome_aba": nome_aba, "prioridade": prioridade})

    if candidatas:
        candidatas = sorted(candidatas, key=lambda item: item["prioridade"])
        return candidatas[0]["nome_aba"], "Aba com data real igual à data selecionada"

    return None, "Nenhuma aba com data real igual à data selecionada encontrada"


def encontrar_aba_por_data_ou_gerencial(wb_origem, data_selecionada, ano_referencia, mes_referencia, permitir_fallback_gerencial=False):
    nome_aba, criterio = encontrar_aba_exata_por_data(
        wb_origem=wb_origem,
        data_selecionada=data_selecionada,
        ano_referencia=ano_referencia,
        mes_referencia=mes_referencia
    )

    if nome_aba:
        return nome_aba, criterio

    if permitir_fallback_gerencial:
        aba_gerencial = encontrar_aba_gerencial(wb_origem)
        if aba_gerencial:
            return aba_gerencial, "Aba Gerencial usada porque nenhuma aba com data real igual à data selecionada foi encontrada"

    return None, "Nenhuma aba com data real igual à data selecionada encontrada"


def extrair_data_nome_consolidado(nome_arquivo):
    nome = str(nome_arquivo)
    padrao = re.search(r"(\d{2})[\.\-/](\d{2})[\.\-/](\d{4})", nome)
    if not padrao:
        return None
    try:
        dia = int(padrao.group(1))
        mes = int(padrao.group(2))
        ano = int(padrao.group(3))
        return datetime(ano, mes, dia).date()
    except ValueError:
        return None


# ============================================================
# CONSOLIDADO EXISTENTE + ARQUIVOS FALTANTES
# ============================================================

def detectar_consolidado_existente(item):
    nome = remover_acentos(item["nome"]).lower()
    if "resumo gerencial cn" in nome or "consolidado" in nome:
        return True

    wb = None
    try:
        wb = load_workbook(item["caminho"], read_only=True, data_only=True, keep_links=False)
        qtd_abas_padrao = sum(1 for aba in wb.sheetnames if identificar_aba_padrao(aba) is not None)
        return qtd_abas_padrao >= 2
    except Exception:
        return False
    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass


def consolidado_serve_para_data(item, data_selecionada, total_datas_selecionadas):
    data_nome = extrair_data_nome_consolidado(item["nome"])
    if data_nome is not None:
        return data_nome == data_selecionada
    return total_datas_selecionadas == 1


def copiar_abas_de_consolidado_existente(wb_origem, wb_final, logs, nome_arquivo, data_selecionada, ano_referencia, mes_referencia):
    copiadas = []

    for nome_aba in wb_origem.sheetnames:
        nome_padrao = identificar_aba_padrao(nome_aba)
        if nome_padrao is None:
            continue

        aba_origem = wb_origem[nome_aba]
        data_real = obter_data_real_da_aba(aba_origem, ano_referencia, mes_referencia)
        if data_real is not None and data_real != data_selecionada:
            logs.append(
                f"Consolidado existente: aba '{nome_aba}' ignorada porque a data real é {data_real.strftime('%d/%m/%Y')}, "
                f"não {data_selecionada.strftime('%d/%m/%Y')}."
            )
            continue

        if aba_padrao_ja_existe(wb_final, nome_padrao):
            logs.append(f"Consolidado existente: aba '{nome_aba}' ignorada porque '{nome_padrao}' já existe no arquivo final.")
            continue

        nome_destino = gerar_nome_aba_unico(nome_padrao, wb_final.sheetnames)
        aba_destino = wb_final.create_sheet(title=nome_destino)
        copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=True)
        limpar_conteudos_a_partir_da_coluna_l(aba_destino)
        aplicar_configuracoes_finais_aba(aba_destino)
        copiadas.append(nome_destino)
        logs.append(f"Consolidado existente '{nome_arquivo}': aba '{nome_aba}' copiada como '{nome_destino}'.")

    return copiadas


# ============================================================
# UPLOAD / SESSÃO
# ============================================================

def criar_pasta_sessao():
    if "pasta_sessao" not in st.session_state:
        st.session_state.pasta_sessao = tempfile.mkdtemp()


def salvar_uploads_em_pasta_sessao(arquivos_upload):
    criar_pasta_sessao()
    arquivos_salvos = []

    for uploaded_file in arquivos_upload:
        nome_limpo = limpar_nome_arquivo(uploaded_file.name)
        caminho_base = os.path.join(st.session_state.pasta_sessao, nome_limpo)
        caminho_final = caminho_base
        contador = 1

        while os.path.exists(caminho_final):
            nome_sem_extensao, extensao = os.path.splitext(nome_limpo)
            caminho_final = os.path.join(
                st.session_state.pasta_sessao,
                f"{nome_sem_extensao}_{contador}{extensao}"
            )
            contador += 1

        with open(caminho_final, "wb") as f:
            f.write(uploaded_file.getvalue())

        arquivos_salvos.append({
            "nome": uploaded_file.name,
            "caminho": caminho_final,
            "tamanho": uploaded_file.size
        })

    return arquivos_salvos


def limpar_arquivos_da_sessao():
    if "pasta_sessao" in st.session_state:
        try:
            if os.path.exists(st.session_state.pasta_sessao):
                shutil.rmtree(st.session_state.pasta_sessao)
        except Exception:
            pass
    st.session_state.arquivos_salvos = []
    st.session_state.uploader_key += 1
    if "pasta_sessao" in st.session_state:
        del st.session_state.pasta_sessao


# ============================================================
# FORMATAÇÃO / LIMPEZA / CÓPIA
# ============================================================

def aplicar_configuracoes_finais_aba(aba_destino):
    aba_destino.sheet_view.zoomScale = 80
    aba_destino.sheet_view.zoomScaleNormal = 80
    aba_destino.freeze_panes = "A5"
    aba_destino.sheet_view.showGridLines = False


def limpar_conteudos_a_partir_da_coluna_l(aba_destino):
    max_linha = aba_destino.max_row
    max_coluna = aba_destino.max_column

    if max_coluna < COLUNA_INICIO_LIMPEZA:
        return

    for intervalo in list(aba_destino.merged_cells.ranges):
        if intervalo.min_col >= COLUNA_INICIO_LIMPEZA:
            celula_principal = aba_destino.cell(row=intervalo.min_row, column=intervalo.min_col)
            if not isinstance(celula_principal, MergedCell):
                celula_principal.value = None
                celula_principal.comment = None
                celula_principal._hyperlink = None

    for linha in aba_destino.iter_rows(
        min_row=1,
        max_row=max_linha,
        min_col=COLUNA_INICIO_LIMPEZA,
        max_col=max_coluna
    ):
        for celula in linha:
            if isinstance(celula, MergedCell):
                continue
            celula.value = None
            celula.comment = None
            celula._hyperlink = None


def garantir_requisitos_todas_abas(wb_final):
    for aba in wb_final.worksheets:
        limpar_conteudos_a_partir_da_coluna_l(aba)
        aplicar_configuracoes_finais_aba(aba)


def padronizar_larguras_colunas_pela_aba_ede(wb_final, nome_aba_referencia="EDE"):
    if nome_aba_referencia not in wb_final.sheetnames:
        return False, f"A aba de referência '{nome_aba_referencia}' não foi encontrada. As larguras das colunas não foram padronizadas."

    aba_referencia = wb_final[nome_aba_referencia]
    max_colunas = max(ws.max_column for ws in wb_final.worksheets)

    for indice_coluna in range(1, max_colunas + 1):
        letra_coluna = aba_referencia.cell(row=1, column=indice_coluna).column_letter
        largura_referencia = aba_referencia.column_dimensions[letra_coluna].width
        if largura_referencia is None:
            largura_referencia = aba_referencia.sheet_format.defaultColWidth or 8.43
        for aba in wb_final.worksheets:
            aba.column_dimensions[letra_coluna].width = largura_referencia

    return True, f"Larguras das colunas padronizadas com base na aba '{nome_aba_referencia}'."


def copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=True):
    for linha in aba_origem.iter_rows(
        min_row=1,
        max_row=aba_origem.max_row,
        min_col=1,
        max_col=aba_origem.max_column
    ):
        for celula_origem in linha:
            celula_destino = aba_destino[celula_origem.coordinate]
            celula_destino.value = celula_origem.value

            if copiar_estilos and celula_origem.has_style:
                celula_destino.font = copy(celula_origem.font)
                celula_destino.fill = copy(celula_origem.fill)
                celula_destino.border = copy(celula_origem.border)
                celula_destino.alignment = copy(celula_origem.alignment)
                celula_destino.number_format = celula_origem.number_format
                celula_destino.protection = copy(celula_origem.protection)

            if celula_origem.comment:
                celula_destino.comment = copy(celula_origem.comment)
            if celula_origem.hyperlink:
                celula_destino._hyperlink = copy(celula_origem.hyperlink)

    for letra_coluna, dimensao_coluna in aba_origem.column_dimensions.items():
        aba_destino.column_dimensions[letra_coluna].width = dimensao_coluna.width
        aba_destino.column_dimensions[letra_coluna].hidden = dimensao_coluna.hidden
        aba_destino.column_dimensions[letra_coluna].outlineLevel = dimensao_coluna.outlineLevel
        aba_destino.column_dimensions[letra_coluna].collapsed = dimensao_coluna.collapsed

    for numero_linha, dimensao_linha in aba_origem.row_dimensions.items():
        aba_destino.row_dimensions[numero_linha].height = dimensao_linha.height
        aba_destino.row_dimensions[numero_linha].hidden = dimensao_linha.hidden
        aba_destino.row_dimensions[numero_linha].outlineLevel = dimensao_linha.outlineLevel
        aba_destino.row_dimensions[numero_linha].collapsed = dimensao_linha.collapsed

    for intervalo_mesclado in aba_origem.merged_cells.ranges:
        aba_destino.merge_cells(str(intervalo_mesclado))

    if aba_origem.auto_filter and aba_origem.auto_filter.ref:
        aba_destino.auto_filter.ref = aba_origem.auto_filter.ref

    aba_destino.freeze_panes = aba_origem.freeze_panes

    try:
        aba_destino.sheet_view.showGridLines = aba_origem.sheet_view.showGridLines
    except Exception:
        pass


def ordenar_abas(wb_final):
    def prioridade(aba):
        nome = aba.title.upper()
        for indice, prefixo in enumerate(ORDEM_ABAS):
            if nome == prefixo or nome.startswith(prefixo + " ") or nome.startswith(prefixo + "_") or nome.startswith(prefixo + "("):
                return indice
        return 99
    wb_final._sheets = sorted(wb_final.worksheets, key=prioridade)


# ============================================================
# PROCESSAMENTO
# ============================================================

def processar_arquivos_salvos(
    arquivos_salvos,
    data_referencia,
    data_selecionada,
    copiar_estilos=True,
    exibir_progresso=True,
    callback_progresso=None,
    total_datas_selecionadas=1,
    permitir_fallback_gerencial=False
):
    logs = []
    arquivos_processados = []
    arquivos_com_erro = []
    abas_criadas = []

    ano_referencia = data_referencia.year
    mes_referencia = data_referencia.month
    nome_arquivo_final = f"Resumo Gerencial CN - {data_selecionada.strftime('%d.%m.%Y')}.xlsx"

    wb_final = Workbook()
    wb_final.remove(wb_final.active)

    logs.append("Iniciando processamento.")
    logs.append(f"Data do consolidado: {data_selecionada.strftime('%d/%m/%Y')}")
    logs.append("Modo seguro: a data é validada pelo conteúdo/cabeçalho da planilha.")

    barra_progresso = st.progress(0) if exibir_progresso else None
    texto_status = st.empty() if exibir_progresso else None
    inicio_local = time.time()

    itens_consolidados = []
    itens_origem = []

    for item in arquivos_salvos:
        if detectar_consolidado_existente(item):
            itens_consolidados.append(item)
        else:
            itens_origem.append(item)

    # 1) Reaproveita somente consolidados da mesma data.
    for item in itens_consolidados:
        wb_consolidado = None
        try:
            if not consolidado_serve_para_data(item, data_selecionada, total_datas_selecionadas):
                logs.append(f"Consolidado existente '{item['nome']}' ignorado para a data {data_selecionada.strftime('%d/%m/%Y')}.")
                continue

            wb_consolidado = load_workbook(item["caminho"], data_only=True, keep_links=False)
            abas_copiadas = copiar_abas_de_consolidado_existente(
                wb_origem=wb_consolidado,
                wb_final=wb_final,
                logs=logs,
                nome_arquivo=item["nome"],
                data_selecionada=data_selecionada,
                ano_referencia=ano_referencia,
                mes_referencia=mes_referencia
            )
            abas_criadas.extend(abas_copiadas)

            if abas_copiadas:
                arquivos_processados.append({
                    "arquivo": item["nome"],
                    "aba_origem": "Consolidado existente da mesma data",
                    "aba_final": ", ".join(abas_copiadas),
                    "criterio": f"Abas reaproveitadas de consolidado existente | Data do consolidado: {data_selecionada.strftime('%d/%m/%Y')}"
                })

        except Exception as erro:
            arquivos_com_erro.append({"arquivo": item["nome"], "erro": str(erro)})
            logs.append(f"ERRO ao reaproveitar consolidado existente {item['nome']} - {erro}")
        finally:
            try:
                if wb_consolidado:
                    wb_consolidado.close()
            except Exception:
                pass

    # 2) Processa arquivos de origem/faltantes, mas SOMENTE se o conteúdo da aba for da data escolhida.
    total_itens_origem = max(len(itens_origem), 1)

    for indice, item in enumerate(itens_origem, start=1):
        nome_original = item["nome"]
        caminho_arquivo = item["caminho"]
        wb_origem = None

        if texto_status:
            decorrido, restante = calcular_estimativa_tempo(inicio_local, indice - 1, total_itens_origem)
            texto_status.info(
                f"Processando {indice}/{len(itens_origem)}: {nome_original} | "
                f"Decorrido: {decorrido} | Tempo restante estimado: {restante}"
            )

        try:
            if not nome_original.lower().endswith(EXTENSOES_EXCEL_VALIDAS):
                raise ValueError("Extensão inválida. Use .xlsx ou .xlsm.")
            if not os.path.exists(caminho_arquivo):
                raise FileNotFoundError("Arquivo temporário não encontrado. Carregue o arquivo novamente.")

            nome_base_personalizado = obter_nome_aba_final_personalizado(nome_original)
            nome_padrao = identificar_aba_padrao(nome_base_personalizado) or nome_base_personalizado

            if aba_padrao_ja_existe(wb_final, nome_padrao):
                logs.append(f"Arquivo '{nome_original}' ignorado porque a aba '{nome_padrao}' já existe no consolidado final.")
                if callback_progresso:
                    callback_progresso(nome_original, data_selecionada)
                continue

            wb_origem = load_workbook(filename=caminho_arquivo, data_only=True, keep_links=False)

            nome_aba_origem, criterio_aba = encontrar_aba_por_data_ou_gerencial(
                wb_origem=wb_origem,
                data_selecionada=data_selecionada,
                ano_referencia=ano_referencia,
                mes_referencia=mes_referencia,
                permitir_fallback_gerencial=permitir_fallback_gerencial
            )

            if not nome_aba_origem:
                logs.append(
                    f"IGNORADO: {nome_original} não possui aba com data real {data_selecionada.strftime('%d/%m/%Y')} no conteúdo/cabeçalho."
                )
                if callback_progresso:
                    callback_progresso(nome_original, data_selecionada)
                continue

            aba_origem = wb_origem[nome_aba_origem]
            data_real_aba = obter_data_real_da_aba(aba_origem, ano_referencia, mes_referencia)
            if data_real_aba is not None and data_real_aba != data_selecionada:
                logs.append(
                    f"IGNORADO: {nome_original} aba '{nome_aba_origem}' tem data real {data_real_aba.strftime('%d/%m/%Y')}, "
                    f"não {data_selecionada.strftime('%d/%m/%Y')}."
                )
                if callback_progresso:
                    callback_progresso(nome_original, data_selecionada)
                continue

            nome_aba_final = gerar_nome_aba_unico(nome_base_personalizado, wb_final.sheetnames)
            aba_destino = wb_final.create_sheet(title=nome_aba_final)

            copiar_aba_como_valores_mais_leve(aba_origem, aba_destino, copiar_estilos=copiar_estilos)
            limpar_conteudos_a_partir_da_coluna_l(aba_destino)
            aplicar_configuracoes_finais_aba(aba_destino)

            arquivos_processados.append({
                "arquivo": nome_original,
                "aba_origem": nome_aba_origem,
                "aba_final": nome_aba_final,
                "criterio": f"{criterio_aba} | Data real validada: {data_selecionada.strftime('%d/%m/%Y')}"
            })
            abas_criadas.append(nome_aba_final)
            logs.append(f"SUCESSO: {nome_original} | aba '{nome_aba_origem}' -> '{nome_aba_final}'.")

        except Exception as erro:
            arquivos_com_erro.append({"arquivo": nome_original, "erro": str(erro)})
            logs.append(f"ERRO: {nome_original} - {erro}")

        finally:
            try:
                if wb_origem:
                    wb_origem.close()
            except Exception:
                pass
            try:
                del wb_origem
            except Exception:
                pass
            gc.collect()
            if barra_progresso:
                barra_progresso.progress(indice / total_itens_origem)
            if callback_progresso:
                callback_progresso(nome_original, data_selecionada)

    if texto_status:
        decorrido, _ = calcular_estimativa_tempo(inicio_local, total_itens_origem, total_itens_origem)
        texto_status.success(f"Processamento concluído. Tempo total: {decorrido}.")

    if len(wb_final.sheetnames) == 0:
        raise Exception(f"Nenhuma aba com data real {data_selecionada.strftime('%d/%m/%Y')} foi criada no arquivo final.")

    garantir_requisitos_todas_abas(wb_final)
    _, mensagem_larguras = padronizar_larguras_colunas_pela_aba_ede(wb_final, "EDE")
    logs.append(mensagem_larguras)
    garantir_requisitos_todas_abas(wb_final)
    ordenar_abas(wb_final)

    caminho_saida_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    wb_final.save(caminho_saida_temp)
    wb_final.close()

    with open(caminho_saida_temp, "rb") as f:
        conteudo_final = f.read()

    try:
        os.remove(caminho_saida_temp)
    except Exception:
        pass

    output = BytesIO(conteudo_final)
    output.seek(0)
    gc.collect()

    return {
        "arquivo_excel": output,
        "nome_arquivo_final": nome_arquivo_final,
        "logs": logs,
        "arquivos_processados": arquivos_processados,
        "arquivos_com_erro": arquivos_com_erro,
        "abas_criadas": abas_criadas,
        "data_processada": data_selecionada
    }


def processar_multiplas_datas(arquivos_salvos, datas_selecionadas, data_referencia, copiar_estilos=True, permitir_fallback_gerencial=False):
    resultados = []
    datas_ordenadas = sorted(datas_selecionadas, key=lambda item: item["data"], reverse=True)

    total_etapas = max(len(datas_ordenadas) * len(arquivos_salvos), 1)
    etapas_concluidas = 0
    inicio_geral = time.time()

    barra = st.progress(0)
    status = st.empty()
    tempo_box = st.empty()

    def atualizar_progresso(nome_arquivo, data_atual):
        nonlocal etapas_concluidas
        etapas_concluidas += 1
        progresso = min(etapas_concluidas / total_etapas, 1)
        barra.progress(progresso)
        decorrido, restante = calcular_estimativa_tempo(inicio_geral, etapas_concluidas, total_etapas)
        tempo_box.info(
            f"⏱️ Tempo decorrido: {decorrido} | "
            f"Tempo restante estimado: {restante} | "
            f"Progresso: {etapas_concluidas}/{total_etapas} etapas"
        )
        status.info(
            f"Processando data {data_atual.strftime('%d/%m/%Y')} | "
            f"Arquivo: {nome_arquivo}"
        )

    for indice, item_data in enumerate(datas_ordenadas, start=1):
        data = item_data["data"]
        status.info(f"Gerando consolidado {indice}/{len(datas_ordenadas)} somente com a data {data.strftime('%d/%m/%Y')}")

        resultado = processar_arquivos_salvos(
            arquivos_salvos=arquivos_salvos,
            data_referencia=data_referencia,
            data_selecionada=data,
            copiar_estilos=copiar_estilos,
            exibir_progresso=False,
            callback_progresso=atualizar_progresso,
            total_datas_selecionadas=len(datas_ordenadas),
            permitir_fallback_gerencial=permitir_fallback_gerencial
        )
        resultados.append(resultado)

    decorrido, _ = calcular_estimativa_tempo(inicio_geral, total_etapas, total_etapas)
    status.success("Todos os consolidados por data foram gerados.")
    tempo_box.success(f"✅ Processamento concluído. Tempo total: {decorrido}.")
    return resultados


def gerar_zip_resultados(resultados):
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for resultado in resultados:
            zip_file.writestr(resultado["nome_arquivo_final"], resultado["arquivo_excel"].getvalue())
    zip_buffer.seek(0)
    return zip_buffer


# ============================================================
# ESTADO DA SESSÃO
# ============================================================

if "arquivos_salvos" not in st.session_state:
    st.session_state.arquivos_salvos = []

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0


# ============================================================
# INTERFACE
# ============================================================

st.title("📊 Sistema de Consolidação Gerencial CN")

st.markdown(
    """
    Este sistema consolida abas de arquivos Excel em **um ou mais arquivos finais separados por data**.

    **Exemplo:** se você selecionar **03/06** e **04/06**, o sistema gera **2 planilhas consolidadas**:
    uma somente para **03/06** e outra somente para **04/06**.

    **Validação de data:** o sistema lê a data no conteúdo da própria planilha, por exemplo: **Data: 3 de junho de 2026**.
    """
)

st.divider()

# Sidebar
st.sidebar.header("⚙️ Configurações")

usar_data_ontem = st.sidebar.checkbox("Usar mês/ano de ontem como referência", value=True)
if usar_data_ontem:
    data_referencia = datetime.today() - timedelta(days=1)
    st.sidebar.info(f"Mês/ano de referência: {data_referencia.strftime('%m/%Y')}")
else:
    data_escolhida = st.sidebar.date_input(
        "Escolha uma data de referência para mês/ano",
        value=(datetime.today() - timedelta(days=1)).date()
    )
    data_referencia = datetime.combine(data_escolhida, datetime.min.time())

qtde_esperada = st.sidebar.number_input("Quantidade esperada de arquivos", min_value=1, max_value=100, value=7, step=1)
copiar_estilos = st.sidebar.checkbox("Copiar formatação das células", value=True)
quantidade_datas_filtro = st.sidebar.number_input("Quantidade máxima de datas no filtro", min_value=1, max_value=31, value=5, step=1)
max_datas_processar = st.sidebar.number_input("Máximo de datas para processar de uma vez", min_value=1, max_value=10, value=5, step=1)
permitir_fallback_gerencial = st.sidebar.checkbox(
    "Permitir usar aba Gerencial quando a aba da data não existir",
    value=False,
    help="Deixe desmarcado para garantir que cada consolidado tenha somente abas com data real igual à data selecionada."
)

st.sidebar.caption("O filtro lê datas dentro das abas e também considera nomes de abas numéricos quando não houver data no conteúdo.")

# Upload
st.subheader("1. Envie os arquivos Excel")

arquivos_upload = st.file_uploader(
    "Selecione todos os arquivos Excel de uma vez. Você pode incluir consolidados já gerados e planilhas faltantes.",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
    key=f"uploader_{st.session_state.uploader_key}"
)

col_carregar, col_limpar = st.columns(2)

with col_carregar:
    if st.button("📥 Carregar arquivos enviados", disabled=not arquivos_upload):
        try:
            limpar_arquivos_da_sessao()
            st.session_state.arquivos_salvos = salvar_uploads_em_pasta_sessao(arquivos_upload)
            st.success(f"{len(st.session_state.arquivos_salvos)} arquivo(s) carregado(s) com sucesso.")
            st.rerun()
        except Exception as erro:
            st.error(f"Erro ao carregar arquivos: {erro}")

with col_limpar:
    if st.button("🧹 Limpar arquivos carregados"):
        limpar_arquivos_da_sessao()
        st.success("Arquivos carregados foram limpos.")
        st.rerun()

arquivos_salvos = st.session_state.arquivos_salvos

if arquivos_salvos:
    st.success(f"{len(arquivos_salvos)} arquivo(s) carregado(s).")

    tamanho_total_mb = sum(item["tamanho"] for item in arquivos_salvos) / (1024 * 1024)
    st.info(f"Tamanho total carregado: {tamanho_total_mb:.2f} MB")

    with st.expander("Ver arquivos carregados"):
        for i, item in enumerate(arquivos_salvos, start=1):
            tamanho_mb = item["tamanho"] / (1024 * 1024)
            nome_padronizado = obter_nome_aba_final_personalizado(item["nome"])
            tipo = "Consolidado existente" if detectar_consolidado_existente(item) else "Planilha de origem/faltante"
            st.write(f"{i}. {item['nome']} — {tamanho_mb:.2f} MB — tipo: {tipo} — aba prevista: {nome_padronizado}")
else:
    st.info("Envie os arquivos Excel e clique em 'Carregar arquivos enviados'.")

st.divider()

# Datas
st.subheader("2. Escolha uma ou mais datas para gerar consolidados separados")

datas_selecionadas = []

if arquivos_salvos:
    datas_para_filtro = listar_datas_disponiveis_arquivos(
        arquivos_salvos=arquivos_salvos,
        ano_referencia=data_referencia.year,
        mes_referencia=data_referencia.month,
        quantidade_datas_filtro=quantidade_datas_filtro
    )

    if datas_para_filtro:
        opcoes_datas = []
        for item in datas_para_filtro:
            data_item = item["data"].date()
            qtd_arquivos = len(set(item["arquivos"]))
            label = f"{data_item.strftime('%d/%m/%Y')} — encontrada em {qtd_arquivos} arquivo(s)"
            opcoes_datas.append({"label": label, "data": data_item})

        datas_selecionadas = st.multiselect(
            "Selecione uma ou mais datas",
            options=opcoes_datas,
            default=opcoes_datas[:1],
            format_func=lambda item: item["label"]
        )

        if datas_selecionadas:
            st.success(f"{len(datas_selecionadas)} data(s) selecionada(s).")
            st.info("Será gerado um arquivo Excel consolidado separado para cada data selecionada.")

            with st.expander("Datas disponíveis"):
                for opcao in opcoes_datas:
                    st.write(f"- {opcao['label']}")
    else:
        st.warning("Nenhuma sequência de datas foi encontrada no mês/ano de referência.")
else:
    st.info("Carregue os arquivos para o sistema listar as datas disponíveis.")

st.divider()

# Processar
st.subheader("3. Processar")

botao_processar = st.button(
    "🚀 Gerar consolidado(s) por data",
    type="primary",
    disabled=len(st.session_state.arquivos_salvos) == 0 or len(datas_selecionadas) == 0
)

if botao_processar:
    try:
        if len(datas_selecionadas) > max_datas_processar:
            st.warning(f"Selecione no máximo {max_datas_processar} data(s) por execução.")
            st.stop()

        with st.spinner("Gerando consolidado(s) separado(s) por data..."):
            resultados = processar_multiplas_datas(
                arquivos_salvos=st.session_state.arquivos_salvos,
                datas_selecionadas=datas_selecionadas,
                data_referencia=data_referencia,
                copiar_estilos=copiar_estilos,
                permitir_fallback_gerencial=permitir_fallback_gerencial
            )

        st.success("Consolidado(s) gerado(s) com sucesso!")

        total_sucesso = sum(len(r["arquivos_processados"]) for r in resultados)
        total_erros = sum(len(r["arquivos_com_erro"]) for r in resultados)

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Arquivos finais gerados", len(resultados))
        with col2:
            st.metric("Abas processadas com sucesso", total_sucesso)
        with col3:
            st.metric("Erros", total_erros)

        st.subheader("4. Downloads")

        if len(resultados) > 1:
            zip_file = gerar_zip_resultados(resultados)
            st.download_button(
                label="⬇️ Baixar todos os consolidados em ZIP",
                data=zip_file,
                file_name="Consolidados_Gerenciais_CN.zip",
                mime="application/zip"
            )

        st.markdown("### Downloads individuais")
        for resultado in resultados:
            st.download_button(
                label=f"⬇️ Baixar {resultado['nome_arquivo_final']}",
                data=resultado["arquivo_excel"],
                file_name=resultado["nome_arquivo_final"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.subheader("5. Resumo")
        for resultado in resultados:
            data_proc = resultado["data_processada"].strftime("%d/%m/%Y")
            with st.expander(f"Resumo do consolidado da data {data_proc}"):
                if resultado["arquivos_processados"]:
                    st.markdown("#### ✅ Arquivos processados")
                    for item in resultado["arquivos_processados"]:
                        st.write(
                            f"- **Arquivo:** {item['arquivo']} | "
                            f"**Aba origem:** {item['aba_origem']} | "
                            f"**Aba final:** {item['aba_final']} | "
                            f"**Critério:** {item['criterio']}"
                        )

                if resultado["arquivos_com_erro"]:
                    st.markdown("#### ❌ Arquivos com erro")
                    for item in resultado["arquivos_com_erro"]:
                        st.error(f"{item['arquivo']}: {item['erro']}")

                st.markdown("#### Logs")
                st.code("\n".join(resultado["logs"]), language="text")

    except Exception as erro:
        st.error(f"Erro geral ao processar os arquivos: {erro}")
        st.code(traceback.format_exc(), language="text")

st.divider()

st.caption(
    "Observação: cada data selecionada gera um arquivo consolidado separado. "
    "O sistema valida a data pelo conteúdo/cabeçalho da aba, por exemplo 'Data: 3 de junho de 2026'. "
    "Se desejar usar a aba Gerencial como fallback, habilite a opção na barra lateral. "
    "Cada consolidado segue as regras de abas, layout, limpeza da coluna L e ordem COB, CUI, EDE, NOB, PVE, SOB, XAM."
)
